import streamlit as st
import pandas as pd
import sqlite3
import os
from io import BytesIO

DB_PATH = 'db.sqlite3'
EXCEL_PATH = 'Discografia casa Vincenzi.xlsx'

# 🧱 Inizializza il database
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS dischi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            autore TEXT,
            album TEXT,
            anno INTEGER,
            genere TEXT,
            formato TEXT
        )
    ''')
    conn.commit()
    conn.close()

def get_excel_download():
    with open(EXCEL_PATH, 'rb') as f:
        excel_bytes = f.read()
    return excel_bytes

# 📥 Importa da Excel con due fogli
def import_excel():
    if os.path.exists(EXCEL_PATH):
        xls = pd.ExcelFile(EXCEL_PATH)
        df_cd = xls.parse('CD')
        df_cd['formato'] = 'CD'

        df_vinili = xls.parse('Vinili')
        df_vinili['formato'] = 'Vinile'

        df = pd.concat([df_cd, df_vinili], ignore_index=True)
        df.columns = [col.lower() for col in df.columns]  # uniforma nomi colonne

        conn = sqlite3.connect(DB_PATH)
        df.to_sql('dischi', conn, if_exists='replace', index=False)
        conn.close()

# ➕ Aggiungi disco
def insert_disco(autore, album, anno, genere, formato):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO dischi (autore, album, anno, genere, formato) VALUES (?, ?, ?, ?, ?)",
                   (autore, album, anno, genere, formato))
    conn.commit()
    conn.close()

# 🔍 Cerca dischi
def search_dischi(query):
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT autore, album, anno, genere, formato FROM dischi WHERE album LIKE ? OR autore LIKE ?", 
                   (f'%{query}%', f'%{query}%'))
    results = cursor.fetchall()
    conn.close()
    return results

# 📋 Mostra tutti i dischi
def get_all_dischi():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT autore, album, anno, genere, formato FROM dischi", conn)
    conn.close()
    return df

# Recupera autori esistenti
def get_autori():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT autore FROM dischi ORDER BY autore")
    autori = [row[0] for row in cursor.fetchall()]
    conn.close()
    return autori

from openpyxl import load_workbook

def append_to_excel(autore, album, anno, genere, formato):
    if not os.path.exists(EXCEL_PATH):
        return  # Se il file non esiste, salta

    wb = load_workbook(EXCEL_PATH)
    sheet_name = 'CD' if formato == 'CD' else 'Vinili'
    ws = wb[sheet_name]

    # Trova la prima riga vuota
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=2, value=autore)
    ws.cell(row=next_row, column=3, value=album)
    ws.cell(row=next_row, column=4, value=anno)
    ws.cell(row=next_row, column=5, value=genere)
    
    wb.save(EXCEL_PATH)

# 🚀 Interfaccia Streamlit
st.set_page_config(page_title="Archivio Dischi", layout="centered")
st.title("🎵 Archivio Dischi Casalingo")

init_db()
import_excel()

menu = st.sidebar.radio("📁 Menu", ["Visualizza", "Cerca", "Aggiungi"])

if menu == "Visualizza":
    st.subheader("📀 Tutti i dischi")
    df = get_all_dischi()
    st.dataframe(df)
    st.download_button(
        label="📥 Scarica Excel aggiornato",
        data=get_excel_download(),
        file_name="archivio_dischi.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

elif menu == "Cerca":
    query = st.text_input("🔍 Cerca per album o autore")
    if query:
        risultati = search_dischi(query)
        if risultati:
            st.table(risultati)
        else:
            st.warning("Nessun disco trovato.")

elif menu == "Aggiungi":
    st.subheader("➕ Aggiungi un nuovo disco")
    with st.form("add_form"):
        autori_esistenti = get_autori()
        autore = st.selectbox("Autore (scegli o scrivi)", options=[""] + autori_esistenti, index=0)
        autore_libero = st.text_input("Oppure inserisci un nuovo autore")
        # Priorità al campo libero se compilato
        autore_finale = autore_libero if autore_libero else autore
        album = st.text_input("Album")
        anno = st.number_input("Anno", min_value=1900, max_value=2100, step=1)
        genere = st.text_input("Genere")
        formato = st.selectbox("Formato", ["CD", "Vinile"])
        submitted = st.form_submit_button("Aggiungi")
        if submitted and autore_finale:
            insert_disco(autore_finale, album, anno, genere, formato)
            append_to_excel(autore_finale, album, anno, genere, formato)
            st.success(f"Disco di {autore_finale} aggiunto con successo!")
    
